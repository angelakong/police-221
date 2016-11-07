#!/usr/bin/python
# Catherine Xu

import random
import collections
import math
import sys
from collections import Counter
from util import *
import openpyxl

def importData():
    wb = openpyxl.load_workbook('project.xlsx')
    allegations_sheet = wb.get_sheet_by_name('Allegations')
    officers_sheet = wb.get_sheet_by_name('Officers')
    complainants_sheet = wb.get_sheet_by_name('Complainants')
    data = list()
    officer = dict()
    complainants = dict()
    pt_ct = 0
    # {'1': ['officer_M', 'officer_White', 'PO'], '2': ['officer_F', 'officer_Hispanic', 'PO']}
    for row in officers_sheet.iter_rows(row_offset=1):
        officer[str(row[0].value)] = ["officer_" + str(row[1].value), "officer_" + str(row[2].value), str(row[3].value)]
    # {'1043921': ['complain_F', 'complain_White/Hispanic'], '1043909': ['complain_F', 'complain_Black']}
    for row in complainants_sheet.iter_rows(row_offset=1):
        complainants[str(row[1].value)] = ["complain_" + str(row[2].value), "complain_" + str(row[5].value)]
    for row in allegations_sheet.iter_rows(row_offset=1):
        if row[1].value != None and row[2].value != None and row[3].value != None and row[6].value != None: 
            all_info = list()
            if str(row[1].value) in complainants and str(row[2].value) in officer:
                complainant_info = complainants[str(row[1].value)]
                officer_info = officer[str(row[2].value)]
                allegation_cat = [str(row[3].value.lower())]
                verdict = str(row[6].value.lower())
                all_info = allegation_cat + officer_info + complainant_info
                if verdict == 'unsustained':
                    data.append((all_info, 0))
                elif verdict == 'sustained':
                    data.append((all_info, 1))
    # (['operation/personnel violations', 'officer_M', 'officer_White', 'PO', 'complain_M', 'complain_Black'], 0)
    return data


def featureExtractor(x):
    features = collections.defaultdict(int)
    for feature in x:
        features[feature] += 1
    return features

def learnPredictor(trainExamples, testExamples, featureExtractor, numIters, eta):
    '''
    Given |trainExamples| and |testExamples| (each one is a list of (x,y)
    pairs), a |featureExtractor| to apply to x, and the number of iterations to
    train |numIters|, the step size |eta|, return the weight vector (sparse
    feature vector) learned.

    You should implement stochastic gradient descent.  

    Note: only use the trainExamples for training!
    You should call evaluatePredictor() on both trainExamples and testExamples
    to see how you're doing as you learn after each iteration.
    '''
    weights = {}
    for x in xrange(numIters):
        for a,b in trainExamples:
            featuresExtracted = featureExtractor(a)
            if (dotProduct(weights, featuresExtracted) * b) < 1:
                increment(weights, b * eta, featuresExtracted)
    return weights  

data = importData()
trainExamples = list() 
testExamples = list()
for i in xrange(300): 
    trainExamples.append(data[i])
# print trainExamples[1:3]
for i in xrange(300, 310):
    testExamples.append(data[i])
weights = learnPredictor(trainExamples, testExamples, featureExtractor, 20, 0.01)  

trainError = evaluatePredictor(trainExamples, lambda(x) : (1 if dotProduct(featureExtractor(x), weights) >= 0 else -1))
devError = evaluatePredictor(testExamples, lambda(x) : (1 if dotProduct(featureExtractor(x), weights) >= 0 else -1))    
print "Official: train error = %s, dev error = %s" % (trainError, devError)

