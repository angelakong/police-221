#!/usr/bin/python

#####################################################
# This file contains the code for implementing      #
# Part I of the CPDP data set.                      #
                                                    #
# Headers marked by the pound sign denote           #
# purpose of the section.                           #
                                                    #
                                                    #
#####################################################

import random
import collections
import math
import sys
import openpyxl
import numpy as np
import matplotlib.pyplot as plt
from collections import Counter
from util import *
from sklearn import svm
from sklearn.feature_extraction import DictVectorizer
from sklearn.model_selection import train_test_split
from sklearn.metrics import accuracy_score
from sklearn.metrics import recall_score
from sklearn.metrics import precision_score
from sklearn import linear_model
from sklearn import tree
from sklearn.cluster import KMeans
from sklearn import datasets
from sklearn.model_selection import cross_val_predict
from sklearn import metrics
from sklearn.metrics import roc_curve, auc, roc_auc_score
from sklearn.metrics import confusion_matrix
from sklearn.feature_selection import RFECV
from sklearn.feature_selection import RFE

allegations = list()
officers = dict()
complainants= dict()
features = DictVectorizer()
results = []

#################################################### REFERENCE VARS
# column numbers for allegation sheet
id_colnum = 0
complainantid_colnum = 1
officerid_colnum = 2
allcat_colnum = 4
result_colnum = 11
investigator_colnum = 22

# column numbers for officer sheet
id_colnum = 0
gender_colnum = 4
race_colnum = 6
rank_colnum = 9
unit_colnum = 10
allegation_count_colnum = 12

# colnum numbers for complainant sheet
cid_colnum = 1
c_gender_colnum = 2
c_age_colnum = 3
c_race_colnum = 5

#################################################### IMPORT DATA
wb = openpyxl.load_workbook('Full Chicago.xlsx')
allegations_sheet = wb.get_sheet_by_name('Allegations')
officers_sheet = wb.get_sheet_by_name('Officers')
complainants_sheet = wb.get_sheet_by_name('Complainants')

# import officer data
for row in officers_sheet.iter_rows(row_offset=1):
    row_dict = {}
    row_dict['officer_gender'] = str(row[gender_colnum].value)
    row_dict['officer_race'] = str(row[race_colnum].value)
    row_dict['officer_rank'] = str(row[rank_colnum].value)
    row_dict['officer_unit'] = str(row[unit_colnum].value) 
    if row[allegation_count_colnum].value is not None:
        row_dict['allegation_count'] = int(row[allegation_count_colnum].value)
    else:
        row_dict['allegation_count'] = 0
    officers[str(row[id_colnum].value)] = row_dict

# import allegation data 
for row in allegations_sheet.iter_rows(row_offset=1):
    row_dict = {}
    row_dict['officer_associated'] = str(row[officerid_colnum].value)
    row_dict['complainant_associated'] = str(row[complainantid_colnum].value)
    row_dict['allegation_category'] = str(row[allcat_colnum].value)
    row_dict['result'] = str(row[result_colnum].value)
    row_dict['investigator'] = str(row[investigator_colnum].value)
    allegations.append(row_dict)

# import complainant data
for row in complainants_sheet.iter_rows(row_offset=1):
    row_dict = {}
    row_dict['complainant_gender'] = str(row[c_gender_colnum].value) 
    row_dict['complainant_age'] = str(row[c_age_colnum].value) 
    row_dict['complainant_race'] = str(row[c_race_colnum].value)
    complainants[str(row[cid_colnum].value)] = row_dict


#################################################### CREATE FEATURE VECTORS
features_temp = list()
pr = 0
for allegation in allegations:
    if allegation['result'] == 'Unsustained':
        result = 0
    elif allegation['result'] == 'Sustained':
        result = 1
    else:
        continue
    feature_vector = collections.defaultdict(int)
    # allegation category 
    feature_vector[allegation['allegation_category']] = 1

    #investigator
    if allegation['investigator'] == 'None':
        feature_vector['investigator'] = 0
    else:
        feature_vector['investigator'] = 1

    #officer information
    associated_officer = officers[allegation['officer_associated']]
    feature_vector['unit'] = associated_officer['officer_unit']
    feature_vector['allegation_count'] = associated_officer['allegation_count']
    #officer race
    feature_vector[associated_officer['officer_race']] = associated_officer['officer_race']
    feature_vector[associated_officer['officer_gender']] = associated_officer['officer_gender']
    feature_vector[associated_officer['officer_rank']] = associated_officer['officer_rank']

    # complainant information -- unused
    # if allegation['complainant_associated'] in complainants:
    #     associated_complainant = complainants[allegation['complainant_associated']]
    #     feature_vector[associated_complainant['complainant_gender']] = associated_complainant['complainant_gender']
    #     feature_vector[associated_complainant['complainant_age']] = associated_complainant['complainant_age']
    #     feature_vector[associated_complainant['complainant_race']] = associated_complainant['complainant_race']

    features_temp.append((feature_vector))
    results.append(result)

#################################################### TRANSFORM FEATURE VECTOR

X = np.array(features.fit_transform(features_temp).toarray())
y = np.array(results)
# print features.feature_names_
# print features.vocabulary_

################################################################################ PERFORM FEATURE SELECTION
# clf = linear_model.LogisticRegression()
# rfe = RFE(clf, step=1)
# rfe.fit(X,y)
# print rfe.ranking_
# print rfe.n_features_
# print rfe.support_
# print("Optimal number of features : %d" % rfecv.n_features_)
# rfecv = RFECV(estimator=clf, step=1, cv=StratifiedKFold(2),
#               scoring='accuracy')# rfecv.fit(X, y)
# rfecv.fit(X,y)
# plt.figure()
# plt.xlabel("Number of features selected")
# plt.ylabel("Cross validation score (nb of correct classifications)")
# plt.plot(range(1, len(rfecv.grid_scores_) + 1), rfecv.grid_scores_)
# plt.show()

################################################################################ STRATEFIED SAMPLING for train/test data sets
# sss = StratifiedShuffleSplit(y, n_iter=3, test_size=0.3, random_state=42)
# X_train, X_test, y_train, y_test = None
# for train_index, test_index in sss:
#     X_train, X_test = X[train_index], X[test_index]
#     y_train, y_test = y[train_index], y[test_index]

################################################################################ SVM MODEL
# clf_svc = svm.SVC()
# clf_svc.fit(X_train, y_train)
# y_pred_train = clf_svc.predict(X_train)
# y_pred = clf_svc.predict(X_test)
# y_score = clf_svc.decision_function(X_test)
# print "SVM accuracy: TRAIN ", accuracy_score(y_train, y_pred_train)
# print "SVM precision: TRAIN", precision_score(y_train, y_pred_train, average='weighted')
# print "SVM recall: TRAIN", recall_score(y_train, y_pred_train, average='weighted')
# print "SVM accuracy: TEST ", accuracy_score(y_test, y_pred)
# print "SVM precision: TEST", precision_score(y_test, y_pred, average='weighted')
# print "SVM recall: TEST", recall_score(y_test, y_pred, average='weighted')

################################################################################ GENERATE ROC CURVE
# fpr, tpr, thresholds = roc_curve(y_test.ravel(), y_score.ravel(), pos_label=1)
# print roc_auc_score(y_test.ravel(), y_score.ravel())

# plt.figure()
# lw = 2
# plt.plot(fpr, tpr, color='darkorange',
#          lw=2, label='ROC curve (area = %0.2f)' % roc_auc_score(y_test.ravel(), y_score.ravel()))
# plt.plot([0, 1], [0, 1], color='navy', lw=lw, linestyle='--')
# plt.xlim([0.0, 1.0])
# plt.ylim([0.0, 1.05])
# plt.xlabel('False Positive Rate')
# plt.ylabel('True Positive Rate')
# plt.title('ROC Curve based on Complaint')
# plt.legend(loc="lower right")
# plt.show()

################################################################################ LOGISTIC REGRESSION MODEL
# clf_logreg = linear_model.LogisticRegression()
# clf_logreg.fit(X_train, y_train)
# y_pred = clf_logreg.predict(X_test)
# y_pred_train = clf_logreg.predict(X_train)
# print len(y_pred), '!!!!!!!!!!!!!!!!!!!!'
# print "LR accuracy: TRAIN ", accuracy_score(y_train, y_pred_train)
# print "LR precision: TRAIN ", precision_score(y_train, y_pred_train, average='weighted')
# print "LR recall: TRAIN ", recall_score(y_train, y_pred_train, average='weighted')
# print "LR accuracy: TEST ", accuracy_score(y_test, y_pred)
# print "LR precision: TEST ", precision_score(y_test, y_pred, average='weighted')
# print "LR recall: TEST ", recall_score(y_test, y_pred, average='weighted')

################################################################################ DECISION TREE MODEL
# clf_tree = tree.DecisionTreeClassifier()
# clf_tree.fit(X_train, y_train)
# y_pred = clf_tree.predict(X_test)
# y_pred_train = clf_tree.predict(X_train)
# # print len(y_pred)
# print "Tree accuracy: TRAIN ", accuracy_score(y_train, y_pred_train)
# print "Tree precision: TRAIN ", precision_score(y_train, y_pred_train, average='weighted')
# print "Tree recall: TRAIN ", recall_score(y_train, y_pred_train, average='weighted')
# print "Tree accuracy: TEST ", accuracy_score(y_test, y_pred)
# print "Tree precision: TEST ", precision_score(y_test, y_pred, average='weighted')
# print "Tree recall: TEST ", recall_score(y_test, y_pred, average='weighted')




